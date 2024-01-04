import re
import time
import uuid
import json
import threading
import os.path
import traceback
from datetime import date, datetime
from itertools import groupby
from collections import OrderedDict
from operator import attrgetter, itemgetter

import pytz
import openpyxl

from django.db.models import Q
from django.db import connection, connections, transaction
from django.conf import settings
from django.views import View
from django.views.static import serve
from django.views.decorators.csrf import csrf_exempt
from django.http import Http404
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.core.exceptions import PermissionDenied
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

from django_celery_beat.models import CrontabSchedule, PeriodicTask, PeriodicTasks

from rest_framework import mixins, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.generics import RetrieveAPIView, ListAPIView, GenericAPIView, ListCreateAPIView
from rest_framework.exceptions import ValidationError
from rest_framework.schemas import AutoSchema, ManualSchema
from rest_framework.schemas.coreapi import coreschema
import coreapi

from fosun_circle.core.ding_talk.open_api import DingTalkMessageOpenApi
from fosun_circle.libs.log import dj_logger as logger
from fosun_circle.core.views import ListVueView, SingleVueView
from fosun_circle.apps.ding_talk.tasks.task_send_ding_message import send_ding_message
from fosun_circle.apps.ding_talk.tasks.task_recall_ding_message import recall_ding_message
from users.models import CircleUsersModel, CircleDepartmentModel
from questionnaire.service import SurveyVoteService
from common.views import CronParseApi

from . import forms
from .models import (
    DingAppTokenModel,
    DingMsgPushLogModel,
    DingMessageModel,
    DingAppMediaModel,
    DingMsgRecallLogModel,
    DingPeriodicTaskModel,
)
from .serializers import (
    DingAppTokenSerializer,
    PushDingMsgLogSerializer,
    ListDingMessageLogSerializer,
    DingAppMediaSerializer,
    DingMessageSerializer,
    ListRecallMsgLogSerializer,
    DingPeriodicTaskSerializer,
)
from fosun_circle.libs.utils.crypto import BaseCipher
from fosun_circle.core.ding_talk.uuc import DingUser


class ObtainUserByCodeApi(APIView):
    throttle_classes = ()

    def post(self, request, *args, **kwargs):
        """ useless """
        code = request.data.get('code')
        user_dict = DingUser(code=code).get_ding_user()
        return Response(data=user_dict)


class ListDingAppTokenApi(ListAPIView):
    queryset = DingAppTokenModel.objects.filter(is_del=False).all()
    serializer_class = DingAppTokenSerializer


class RetrieveDingAppTokenApi(RetrieveAPIView):
    queryset = DingAppTokenModel.objects.filter(is_del=False).all()
    serializer_class = DingAppTokenSerializer


class RetrieveDingAppTokenByAgentIdApi(RetrieveAPIView):
    lookup_field = "agent_id"

    queryset = DingAppTokenModel.objects.filter(is_del=False).all()
    serializer_class = DingAppTokenSerializer


class CreateDingAppTokenApi(mixins.CreateModelMixin,
                            mixins.UpdateModelMixin,
                            GenericAPIView):
    serializer_class = DingAppTokenSerializer

    def get_object(self):
        agent_id = self.request.data["agent_id"]
        return DingAppTokenModel.objects.get(agent_id=agent_id)

    def post(self, request, *args, **kwargs):
        """ 添加或更新微信用 """
        agent_id = request.data["agent_id"]
        app_obj = DingAppTokenModel.objects.filter(agent_id=agent_id).first()

        if not app_obj:
            return self.create(request, *args, **kwargs)

        return self.update(request, *args, **kwargs)


class CreateDingMessageView(SingleVueView):
    template_name = "ding_talk/ding_message_create.html"
    model = DingMessageModel

    def get_queryset(self):
        mobile = self.request.user.mobile
        user = CircleUsersModel.get_user_by_mobile(mobile, add_uuc=False)

        log_args = (mobile, user.is_superuser, user.is_vote_admin)
        logger.info("CreateDingMessageView.user -> mobile:%s, is_superuser: %s, is_vote_admin: %s", *log_args)

        app_queryset = DingAppTokenModel.objects.filter(is_del=False).values("id", "app_name", "app_token")
        msg_type_list = self.get_msg_type_list()

        # 部门下拉列表
        department_tree = CircleDepartmentModel.get_department_tree("root", on_cascade=True)
        department_list = [dict(label="全员", value=department_tree["value"])]
        department_list.extend(department_tree.get("children", []))

        return dict(
            app_list=list(app_queryset), msg_type_list=msg_type_list,
            source_list=self.get_source_list(), department_list=department_list,
            app_media_mapping=self.get_media_map(),  # 应用下媒体文件
        )

    @classmethod
    def get_media_map(cls):
        app_media_mapping = {}
        media_queryset = DingAppMediaModel.objects.filter(is_success=True)

        for app_id, iterator in groupby(media_queryset, key=attrgetter("app_id")):
            media_list = app_media_mapping.setdefault(app_id,  [])

            for obj in iterator:
                media_list.append(
                    dict(
                        app_id=app_id, media_title=obj.media_title,
                        media_type=obj.media_type, media_id=obj.media_id
                    )
                )

        return app_media_mapping

    @classmethod
    def get_msg_type_list(cls, is_all=False):
        return [
            dict(msg_type=item[0], msg_type_cn=item[1])
            for item in DingMessageModel.MSG_TYPE_CHOICES
            if item[1] == 'oa'
        ]

    @classmethod
    def get_source_list(cls):
        return [
            dict(source=item[0], source_cn=item[1])
            for item in cls.model.SOURCE_CHOICES
        ]


class ListDingMessageLogView(ListVueView):
    template_name = "ding_talk/ding_message_list.html"
    serializer_class = ListDingMessageLogSerializer

    def get_pagination_list(self):
        # mobile = self.request.user.mobile
        # user = CircleUsersModel.get_user_by_mobile(mobile, add_uuc=False)

        message_queryset = DingMsgPushLogModel.objects.filter(is_del=False)
        serializer = self.serializer_class(message_queryset[:10], many=True)

        return dict(list=serializer.data, total_count=message_queryset.count())


class ListDingMessageMediaView(ListVueView):
    template_name = "ding_talk/ding_message_media.html"
    serializer_class = DingAppMediaSerializer

    def get_pagination_list(self):
        app_queryset = DingAppTokenModel.objects.filter(is_del=False).values("id", "app_name", "app_token")
        media_list = [dict(type_id=k, media_type=v) for k, v in DingAppMediaModel.MEDIA_TYPE_CHOICE]

        queryset = DingAppMediaModel.objects.filter(is_success=True).all()
        serializer = self.serializer_class(queryset[:10], many=True)

        return dict(
            app_list=list(app_queryset), media_list=media_list,
            list=serializer.data, total_count=queryset.count()
        )


class ListDingPushMsgLogApi(ListAPIView):
    # STOP_PAGINATOR = True
    serializer_class = ListDingMessageLogSerializer

    def execute_sql(self, sql, params=None, many=True):
        cursor = connection.cursor()
        cursor.execute(sql, params=params)

        if many:
            return cursor.fetchall()

        return cursor.fetchone()

    def get_raw_queryset(self, page, page_size=10):
        # Perform the lookup filtering.
        # lookup_kwarg = self.kwargs
        query_params = self.request.query_params

        msg_model = DingMessageModel
        log_model = DingMsgPushLogModel

        fields = ['a.%s' % col_name for col_name in log_model.fields()]
        sql = "SELECT %s FROM %s a JOIN %s b ON a.ding_msg_id=b.id WHERE a.is_del=false AND b.is_del=false "
        sql_params = [log_model._meta.db_table, msg_model._meta.db_table]

        if query_params.get("app_id"):
            sql += " AND b.app_id=%s "
            sql_params.append(query_params["app_id"])

        if query_params.get("msg_type"):
            sql += " AND b.msg_type=%s "
            sql_params.append(query_params["msg_type"])

        if query_params.get("mobiles"):
            mobiles = query_params["mobiles"]
            mobile_list = ["'%s'" % m.strip() for m in mobiles.split(",") if m.strip()]

            sql += " AND a.receiver_mobile IN (%s) "
            sql_params.append(", ".join(mobile_list))

        if query_params.get("msg_title"):
            sql += " AND b.msg_title LIKE '%s' "
            sql_params.append('%%%%%s%%%%' % query_params["msg_title"])

        # Calculate Count
        db_ret = self.execute_sql(sql % tuple(["COUNT(a.id)"] + sql_params))
        total_count = db_ret[0][0] if db_ret else 0

        # Paginate
        sql += " ORDER BY id DESC LIMIT %s OFFSET %s" % (page_size, (page - 1) * page_size)

        sql_params = tuple([", ".join(fields)] + sql_params)
        logger.info("ListDingPushMsgLogApi.get_queryset SQL: %s", sql % sql_params)
        return log_model.objects.raw(sql % sql_params), total_count

    def get(self, request, *args, **kwargs):
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))

        queryset, total_count = self.get_raw_queryset(page, page_size)
        serializer = self.get_serializer(queryset, many=True)

        data = dict(
            total_count=total_count, page=page,
            pageSize=page_size, list=serializer.data
        )
        return Response(data=data)


class DetailDingPushMsgLogApi(RetrieveAPIView):
    serializer_class = PushDingMsgLogSerializer
    queryset = DingMsgPushLogModel.objects.filter(is_del=False)


@method_decorator(csrf_exempt, name="dispatch")
class PreviewBucketMediaApi(View):
    LOGIN_REQUIRED = False

    def get(self, request, *args, **kwargs):
        """ 文件预览 """
        key_name = kwargs["key_name"]
        access_token = request.GET.get("access_token")

        key, ext = os.path.splitext(key_name)
        media_obj = DingAppMediaModel.get_media_by_key(key=key)

        if not media_obj:
            return Http404()

        if not media_obj.is_share and access_token != media_obj.access_token:
            raise PermissionDenied(403, "您没有权限访问")

        document_root, path = os.path.split(media_obj.media.path)
        return serve(request, path, document_root)


class ListDingMessageInfoApi(APIView):
    def query_queryset(self, query_params=None):
        query_params = query_params or self.request.query_params

        q = Q()
        q.connector = "AND"
        q.children.append(("is_del", False))

        if query_params.get("app_id"):
            q.children.append(("app_id", query_params["app_id"]))

        if query_params.get("msg_title"):
            q.children.append(("msg_title__icontains", query_params["msg_title"]))

        if query_params.get("msg_text"):
            q.children.append(("msg_text__icontains", query_params["msg_text"]))

        if query_params.get('source'):
            q.children.append(("source", query_params["source"]))

        return DingMessageModel.objects.filter(q).all()

    def get(self, request, *args, **kwargs):
        serializer_class = DingMessageSerializer

        query_params = self.request.query_params
        page = int(query_params.get("page") or 1)
        page_size = int(query_params.get("page_size") or 10)

        queryset = self.query_queryset(query_params)[(page - 1) * page_size: page * page_size]
        serializer = serializer_class(queryset, many=True)

        return Response(data=dict(
            list=serializer.data,
            page=int(self.request.query_params.get("page") or 1),
            total_count=self.query_queryset(query_params).count()
        ))


class UploadDingMessageMediaApi(APIView):
    def post(self, request, *args, **kwargs):
        """ 上传钉钉消息媒体文件 """
        data = request.data
        app_token = data.pop("app_token")[0]
        app_obj = DingAppTokenModel.get_app_by_token(app_token=app_token)

        media_type_cn = data.pop("media_type")[0]  # QueryDict => [image], [file], [voice]
        media_mapping = {v: k for k, v in DingAppMediaModel.MEDIA_TYPE_CHOICE}
        media_type = media_mapping[media_type_cn]

        form_data = dict(data, media_title=data.get("media_title", ""), media_type=media_type, app=app_obj.id)
        form = forms.UploadDingMediaForm(form_data, request.FILES)
        if form.is_valid():
            with transaction.atomic():
                media_obj = form.save()
                # media_obj = DingAppMediaModel.objects.create(**form.cleaned_data)

                file_obj = request.FILES["media"]
                ding_service = DingTalkMessageOpenApi(
                    corp_id=app_obj.corp_id, app_key=app_obj.app_key,
                    app_secret=app_obj.app_secret, agent_id=app_obj.agent_id
                )

                if settings.DEBUG:
                    resp = dict(media_id="@debug_%s" % datetime.now().timestamp())
                else:
                    resp = ding_service.upload_media_file(media_type_cn, media_file=file_obj)
                logger.info("UploadDingMessageMediaApi.upload_media_file debug:%s, ret: %s", settings.DEBUG, resp)

                media_obj.media_id = resp["media_id"]
                media_obj.save()

                return Response(data=media_obj.to_dict(exclude=("media",)), status=status.HTTP_200_OK)

        return Response(data=None, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SuccessDingMessageMediaApi(APIView):
    def post(self, request, *args, **kwargs):
        """ 确定媒体文件成功 """
        data = request.data
        media_pk = data.get("media_pk", 0)
        is_success = data.get("is_success", False)

        DingAppMediaModel.objects.filter(id=media_pk).update(is_success=is_success, is_del=not is_success)
        return Response(data=None, status=status.HTTP_200_OK)


class ListDingMessageMediaApi(ListAPIView):
    serializer_class = DingAppMediaSerializer

    def get_queryset(self):
        query_params = self.request.query_params
        app_id = query_params.get("app_id")
        media_title = query_params.get("media_title")
        media_type = query_params.get("media_type")

        query_kwargs = dict(is_success=True)
        app_id and query_kwargs.update(app_id=app_id)
        media_type and query_kwargs.update(media_type=media_type)
        media_title and query_kwargs.update(media_title__contains=media_title)

        return DingAppMediaModel.objects.filter(**query_kwargs).all()


class SendDingAppMessageApi(GenericAPIView):
    throttle_classes = []
    schema = AutoSchema(
        manual_fields=[
            coreapi.Field(name='agent_id', required=True, location='form', description='应用id', type='integer'),

        ])

    MSG_MAX_MQ_SIZE = 50
    MSG_MAX_BULK_SIZE = 1000
    serializer_class = PushDingMsgLogSerializer

    @staticmethod
    def get_mobiles_from_excel(path):
        mobile_list = []

        if not path or not os.path.exists(path):
            return mobile_list

        wb = openpyxl.load_workbook(path)
        sheet = wb.worksheets[0]

        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row, 1).value
            cell_val = str(value or "").strip()

            if cell_val:
                mobile_list.append(cell_val)

        return mobile_list

    def async_send_messages(self, data):
        """ 发送钉钉消息: 先入库, 再通过 mq 发送消息

        :param data: dict or list of dictionary
        :return:
        """
        data.pop("dep_ids", None)
        "is_test" not in data and data.update(is_test=False)

        # 投递到异步任务参数列表的消息最大值
        max_mq_delivery_size = data.pop("mq_delivery_size", self.MSG_MAX_MQ_SIZE)

        # (1) 消息先入库
        serializer = self.serializer_class(data=data)
        serializer.is_valid(raise_exception=True)
        message_body = dict(**data)

        # 手机号拆分, 判断是否批量发送
        receiver_mobile = message_body.pop("receiver_mobile", "")
        mobile_list = [m.strip() for m in receiver_mobile.split(",") if m.strip()]
        many = True if len(mobile_list) > 1 else False

        if many:
            if not mobile_list:
                raise ValueError("参数<receiver_mobile>不能为空")

            if len(mobile_list) > self.MSG_MAX_BULK_SIZE:
                raise ValidationError("手机号数量超过最大限制(500)")

            data_or_list = [dict(**message_body, receiver_mobile=mobile) for mobile in mobile_list]
        else:
            data_or_list = dict(message_body, receiver_mobile=receiver_mobile)

        # many=True 支持批量创建，如果没有在 list_serializer_class 序列化类中重写 create 的批量创建方法,
        # 仅使用 serializer_class 序列化器，会调用 serializer_class.create 方法一个一个来创建，效率偏低
        serializer = self.serializer_class(data=data_or_list, many=many)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        instance_list = instance if isinstance(instance, list) else [instance]

        # (2) 将消息推入MQ中, 异步推送消息
        for i in range(0, len(instance_list), max_mq_delivery_size):
            slice_instances = instance_list[i: i + max_mq_delivery_size]
            msg_uid_list = [msg_obj.msg_uid for msg_obj in slice_instances]

            if not data.get("is_test"):
                send_ding_message.delay(msg_uid_list=msg_uid_list)
            else:
                send_ding_message(msg_uid_list=msg_uid_list)

        return instance_list

    def post(self, request, *args, **kwargs):
        """ 根据手机号推送钉钉消息
        request.data:
            app_token: string, app_token attribute of DingMsgPushLogModel's instance
            msg_title: string, message title
            msg_media: string, message media, like image, file, video. eg: @lADOADmaWMzazQKA
            msg_type： int, look up `DingMsgTypeEnum`
            msg_text: string, message body
            receiver_mobile: string, receiver's mobile to send message, eg: '13600000000,13500000001'
            receiver_job_code: string, [Deprecated] ding's userId to send message eg: '1602133682287,1635343667135'
            source: int, which invoker
            msg_url: string, jump web page from app
            msg_pc_url: string, jump web page from pc
            is_test: bool, default: False, whether to test messages
        """
        data = dict(request.data)
        receiver_mobile = data.pop("receiver_mobile", "")
        mobile_list = [m.strip() for m in receiver_mobile.split(",") if m.strip()]

        # xlsx文件上传
        excel_path = data.pop("batch_mobile_path", None)
        mobile_list.extend(SendDingAppMessageApi.get_mobiles_from_excel(excel_path))
        data.update(receiver_mobile=",".join(mobile_list))

        self.async_send_messages(data=data)
        return Response(data=None, status=status.HTTP_201_CREATED)


class SendDingMessageByDepartmentApi(APIView):
    throttle_classes = []
    MSG_MAX_BULK_SIZE = SendDingAppMessageApi.MSG_MAX_BULK_SIZE

    def bulk_send_ding_messages(self, msg_body, mobile_list):
        """ 子线程发送钉钉消息 """
        total_cnt = len(mobile_list)
        # 每 step_size 个手机号切分一次
        step_size = self.MSG_MAX_BULK_SIZE
        send_ding_api = SendDingAppMessageApi()

        for i in range(0, len(mobile_list), step_size):
            start = i
            stop = i + step_size
            tt = time.time()
            receiver_mobiles = mobile_list[start:stop]
            push_cnt = len(receiver_mobiles)

            log_args = (start, stop, push_cnt, total_cnt)
            logger.info("bulk_send_ding_messages => start: %s, stop: %s, PushCnt: %s, TotalCnt: %s", *log_args)

            try:
                msg_data = dict(msg_body, receiver_mobile=",".join(receiver_mobiles))
                send_ding_api.async_send_messages(data=msg_data)
            except Exception as e:
                logger.info("----- bulk_send_ding_messages error -----")
                logger.error(traceback.format_exc())

            logger.info("bulk_send_ding_messages => PushCnt: %s CostTime: %s", push_cnt, time.time() - tt)

    def post(self, request, *args, **kwargs):
        """ 根据钉钉部门 + 人员, 推送钉钉消息

        request.data:
            dep_ids: list, 部门id, eg: ['c7d6b456-24a5-4a30-973e-9607b0a05071', '9ad3bc09-56e2-4a47-a599-2b11e07768b9']
            app_token: string, app_token attribute of DingMsgPushLogModel's instance
            msg_title: string, message title
            msg_media: string, message media, like image, file, video. eg: @lADOADmaWMzazQKA
            msg_type： int, look up `DingMsgTypeEnum`
            msg_text: string, message body
            receiver_mobile: string, receiver's mobile to send message, eg: '13600000000,13500000001'
            source: int, which invoker
            is_test: bool, default: False, whether to test messages
        """
        data = dict(**request.data)
        mobile_list = []
        dep_ids = data.pop("dep_ids", [])
        logger.info("SendDingMessageByDepartmentApi.dep_ids:%s", dep_ids)

        # 获取该部门和所有子部门的人员手机号
        if dep_ids:
            user_queryset = CircleUsersModel.get_ding_users_by_dep_ids(dep_ids)
            mobile_list = [item["phone_number"] for item in user_queryset]

        # 获取手机号
        receiver_mobile = data.pop("receiver_mobile", "")
        user_mobile_list = [m.strip() for m in receiver_mobile.split(",") if m.strip()]
        mobile_list.extend(user_mobile_list)

        # xlsx文件上传
        excel_path = data.pop("batch_mobile_path", None)
        mobile_list.extend(SendDingAppMessageApi.get_mobiles_from_excel(excel_path))

        if not mobile_list:
            raise ValidationError("未选择任何部门或人员，发送失败")

        if settings.IS_DOCKER:
            # Docker中线程未执行完就结束了(探究: 与 uwsgi 参数 harakiri 有关)
            self.bulk_send_ding_messages(data, mobile_list)
        else:
            # 在ECS等机器中均能正常执行子线程，但Docker中会被突然中断
            t = threading.Thread(target=self.bulk_send_ding_messages, args=(data, mobile_list))
            t.setDaemon(True)
            t.start()

        return Response(data=None, status=status.HTTP_201_CREATED)


class RecallDingAppMsgTaskApi(APIView):
    def post(self, request, *args, **kwargs):
        """ 撤回钉钉推送的消息 """
        # lookup_kwarg = self.kwargs
        data = request.data
        task_id = data.get("task_id") or None
        msg_uid = data.get("msg_uid") or None
        app_token = data.get("app_token")

        app_obj = DingAppTokenModel.get_app_by_token(app_token=app_token)
        log_queryset = DingMsgPushLogModel.objects.all()

        if not task_id and not msg_uid:
            return Response(data=None, status=200)

        if task_id:
            log_queryset = log_queryset.filter(task_id=task_id).all()

        if msg_uid:
            log_queryset = log_queryset.filter(msg_uid=msg_uid).all()

        msg_task_ids = {int(obj.task_id) for obj in log_queryset}

        for task_id in msg_task_ids:
            logger.info("RecallDingAppMsgTaskApi.post => 钉钉消息撤回 OK, task_id<%s>, msg_uid: %s", task_id, msg_uid)
            recall_ding_message.delay(task_id=task_id, app_id=app_obj.id, msg_uid=msg_uid)

        return Response(data=None, status=200)


class AlertAgainDingAppMsgApi(APIView):
    def get_bulk_msg_uid_list(self, ding_msg_id, is_done=False):
        bulk_msg_uid_list = []

        queryset = DingMsgPushLogModel.objects \
            .filter(ding_msg_id=ding_msg_id, is_del=False, is_done=is_done, ) \
            .values_list("msg_uid", flat=True)

        max_step_size = 50
        slice_msg_uid_list = []

        for msg_uid in queryset:
            slice_msg_uid_list.append(msg_uid)

            if len(slice_msg_uid_list) == max_step_size:
                bulk_msg_uid_list.append(slice_msg_uid_list[:])
                slice_msg_uid_list = []
        else:
            # 最后一次可能没有 max_step
            bulk_msg_uid_list.append(slice_msg_uid_list)

        return bulk_msg_uid_list

    def post(self, request, *args, **kwargs):
        """ 钉钉消息二次提醒 """
        data = request.data
        msg_uid = data.get("msg_uid")
        ding_msg_id = data.get("ding_msg_id", 0)
        ihcm_survey_id = data.get("ihcm_survey_id", 0)

        count = 0
        bulk_msg_uid_list = []

        if ihcm_survey_id:
            # 批量二次提醒
            ding_msg_objs = DingMessageModel.objects.filter(ihcm_survey_id=ihcm_survey_id).all()  # 确定只有一个
            if len(ding_msg_objs) > 1:
                raise ValidationError("问卷记录有多条")

            bulk_msg_uid_list.extend(self.get_bulk_msg_uid_list(ding_msg_id=ding_msg_objs[0].id))

        elif ding_msg_id:
            # 话题标签或其他
            bulk_msg_uid_list.extend(self.get_bulk_msg_uid_list(ding_msg_id=ding_msg_id, is_done=False))

        else:
            # 单次提醒(问卷或其他消息)
            if not msg_uid:
                raise ValidationError("钉钉消息<msg_uid>不能为空")

            logger.info("钉钉二次提醒 msg_uid: %s", msg_uid)
            bulk_msg_uid_list.append([msg_uid])

        for msg_uid_list in bulk_msg_uid_list:
            count += len(msg_uid_list)
            logger.info("钉钉消息二次提醒 => count:%s" % count)

            if msg_uid_list:
                send_ding_message.delay(msg_uid_list=msg_uid_list, alert=True)

        return Response(data=None, status=200)


class DetailDingAppMsgApi(APIView):
    def get(self, request, *args, **kwargs):
        """ 钉钉消息详情 """
        ding_msg_id = request.query_params.get("ding_msg_id", 0)
        msg_obj = DingMessageModel.objects.get(id=ding_msg_id, is_del=False)

        media_obj = DingAppMediaModel.objects.filter(media_id=msg_obj.msg_media, is_del=False).first()

        # 序列化
        msg_serializer = DingMessageSerializer(msg_obj)
        media_serializer = DingAppMediaSerializer(media_obj)

        return Response(
            data=dict(msg=msg_serializer.data, media=media_serializer.data),
            status=status.HTTP_200_OK
        )


class ListIhcmSurveyApi(APIView):
    def get(self, request, *args, **kwargs):
        mobile = self.request.user.mobile
        user = CircleUsersModel.get_user_by_mobile(mobile, add_uuc=False)

        # 问卷列表
        vote_service = SurveyVoteService()
        survey_result = vote_service.get_survey_list(dict(size=1000), ding_survey_ids=())
        survey_list = vote_service.get_survey_list_by_auth(survey_result.pop("list", []), user, login_mobile=mobile)

        return Response(data=dict(list=survey_list))


class ListCircleTopicApi(APIView):
    def get(self, request, *args, **kwargs):
        data = []
        conn = connections["bbs_user"]
        cursor = conn.cursor()
        new_line_regex = re.compile(r"\</br\>|\<br/\>", re.S | re.M)

        # 活动贴正常在下拉菜单显示，不管是否被隐藏
        sql = """
            SELECT id, content  FROM "starCircle_starcircle"
            WHERE is_delete=false AND user_id IN (
                SELECT id FROM users_userinfo WHERE fullname LIKE '%星小圈%'
            )
            ORDER BY id DESC
            LIMIT 50
        """
        cursor.execute(sql)
        db_topic_ret = cursor.fetchall()

        for item in db_topic_ret:
            topic_id = item[0]
            text_list = [s.strip() for s in new_line_regex.split(item[1] or "") if s.strip()]

            if not text_list:
                continue

            raw_title = text_list[0]
            title = raw_title[:60] if len(raw_title) > 60 else raw_title
            content = text_list[0] if len(text_list) == 1 else "\n".join(text_list[1:])
            data.append(dict(id=topic_id, title=title, content=content))

        return Response(data=dict(list=data))


class FormDataMessageApi(APIView):
    def get(self, request, *args, **kwargs):
        app_queryset = DingAppTokenModel.objects.filter(is_del=False).values("id", "app_name", "app_token")
        msg_type_list = CreateDingMessageView.get_msg_type_list(is_all=True)

        # 消息来源
        source_list = [dict(source=item[0], source_cn=item[1]) for item in DingMessageModel.SOURCE_CHOICES]
        topic_regex = re.compile(r"/exerland/bbs/filterFind/\d+?/(.*?)$")  # 相关问卷与话题

        survey_list, sent_topic_dict = [], {}
        ding_query = dict(source__in=[2, 3], is_del=False)
        ding_fields = ["msg_title", "msg_url", "id", "source"]
        ding_queryset = DingMessageModel.objects.filter(**ding_query).values(*ding_fields).order_by("source")

        # 数据太多，逻辑有问题
        # for source, ding_iter in groupby(ding_queryset, key=itemgetter("source")):
        #     if source == 2:
        #         survey_list.extend([
        #             dict(source=source, title=item["msg_title"], ding_msg_id=item["id"])
        #             for item in ding_iter
        #         ])
        #     else:
        #         for ding_item in ding_iter:
        #             ding_msg_id = ding_item["id"]
        #             msg_url = ding_item["msg_url"]
        #             match = topic_regex.search(msg_url)
        #
        #             if not match: continue
        #
        #             topic_title = match.group(1)
        #             topic_item = dict(source=source, ding_msg_id=ding_msg_id, title=unquote_plus(topic_title))
        #             sent_topic_dict[ding_msg_id] = topic_item

        return Response(data=dict(
            app_list=list(app_queryset), msg_type_list=msg_type_list, survey_list=survey_list,
            source_list=source_list, sent_topic_list=list(sent_topic_dict.values())
        ))


class BatchMobileExcelUploadApi(APIView):
    def post(self, request, *args, **kwargs):
        upload_file = request.FILES["batch_mobiles"]

        dt_time = datetime.now().strftime("%Y%d%m%H%M%S")
        path = os.path.join(settings.MEDIA_ROOT, str(date.today()))
        uuid_name = str(uuid.uuid1()) + "_" + dt_time + os.path.splitext(upload_file.name)[1]

        if not os.path.exists(path):
            os.makedirs(path)

        save_path = os.path.join(path, uuid_name)
        default_storage.save(save_path, ContentFile(upload_file.read()))  # 保存到本地

        return Response(data=dict(path=save_path))


class RecallMsgLogView(ListVueView):
    template_name = "ding_talk/ding_recall_log_list.html"
    serializer_class = ListRecallMsgLogSerializer

    def get_pagination_list(self):
        app_queryset = DingAppTokenModel.objects.filter(is_del=False).values("id", "app_name", "app_token")
        recall_log_queryset = DingMsgRecallLogModel.objects.filter(is_del=False).all()
        serializer = self.serializer_class(recall_log_queryset[:10], many=True)

        return dict(
            app_list=list(app_queryset),
            list=serializer.data,
            total_count=recall_log_queryset.count(),
        )


class RecallMsgLogApi(ListCreateAPIView):
    serializer_class = ListRecallMsgLogSerializer

    def _get_sql_raw_query(self, model, query_kwargs=None):
        """ 已撤回的消息搜索和撤回弹框搜索 """
        logger.info("RecallMsgLogApi._get_sql_raw_query => db_table: %s, query_kwargs: %s", model, query_kwargs)
        params = query_kwargs or self.request.query_params

        sql = """
            SELECT %s FROM %s a 
            JOIN circle_ding_message_info c 
            ON a.ding_msg_id=c.id 
            WHERE a.is_del=false AND c.is_del=false 
        """
        fields = model.fields()
        field_str = ["a.%s" % col_name for col_name in fields]
        sql_params = [", ".join(field_str), model._meta.db_table]

        if params.get("app_id"):
            if 'app_id' in fields:
                sql += " AND a.app_id IN (SELECT id FROM circle_ding_app_token WHERE id=%s AND is_del=false) "
            else:
                sql += """
                    AND a.ding_msg_id IN (
                        SELECT aa.id FROM circle_ding_message_info aa 
                        JOIN circle_ding_app_token bb 
                        ON aa.app_id=bb.id
                        WHERE bb.id=%s AND aa.is_del=false AND bb.is_del=false 
                    ) 
                """

            sql_params.append(params['app_id'])

        if params.get("task_id"):
            sql += " AND a.task_id='%s' "
            sql_params.append(params['task_id'])

        if params.get("msg_title"):
            sql += " AND c.msg_title LIKE '%s' "
            sql_params.append('%%%%%s%%%%' % params['msg_title'])

        if 'is_recall' in fields:
            sql += " AND a.is_recall=false "

        if params.get('time_range'):
            sql += " AND a.receive_time >='%s' AND a.receive_time <='%s' "
            sql_params.extend(params['time_range'])

        sql += ' ORDER BY id DESC'
        sql_params = tuple(sql_params)
        logger.info("RecallMsgLogApi._get_sql_raw_query SQL: %s", sql % sql_params)
        return sql % sql_params

    def get_queryset(self):
        """ 回撤消息的列表接口 """
        query_params = self.request.query_params
        sql = self._get_sql_raw_query(DingMsgRecallLogModel, query_params)

        return DingMsgRecallLogModel.objects.raw(sql)

    def deliver_recall_to_mq(self, task_id_list=None, app_id=None):
        task_id_list = task_id_list or []

        for task_id in task_id_list:
            if not task_id:
                continue

            try:
                recall_ding_message.delay(task_id=task_id, app_id=app_id)
                logger.info("RecallMsgLogApi._deliver_recall_to_mq => 消息撤回, task_id: %s", ",".join(task_id_list))
            except Exception as e:
                logger.error("RecallMsgLogApi._deliver_recall_to_mq => 消息撤回 err: %s", e)
                logger.error(traceback.format_exc())

    def post(self, request, *args, **kwargs):
        """ 回撤消息的提交接口 """
        data = self.request.data
        app_id = data["app_id"]
        logger.info("RecallMsgLogApi.post => 消息撤回 data: %s", data)

        sql = self._get_sql_raw_query(DingMsgPushLogModel, data)
        raw_queryset = DingMsgPushLogModel.objects.raw(sql)
        task_id_list = list(set([obj.task_id for obj in raw_queryset]))

        t_kwargs = dict(task_id_list=task_id_list, app_id=app_id)
        t = threading.Thread(target=self.deliver_recall_to_mq, kwargs=t_kwargs)
        t.setDaemon(True)
        t.start()

        return Response(data=None)


class SelectDepartmentApi(APIView):
    def get(self, request, *args, **kwargs):
        """ 部门下拉列表 """
        department_tree = CircleDepartmentModel.get_department_tree("root", on_cascade=True)
        department_list = [dict(label="全员", value=department_tree["value"])]
        department_list.extend(department_tree.get("children", []))

        return Response(data=dict(list=department_list))


class AddPeriodicTaskView(SingleVueView):
    """ 添加定时任务 """
    template_name = "ding_talk/ding_periodic_task_add.html"

    @classmethod
    def get_message_queryset(cls, params=None):
        params = params or {}
        q1 = Q(("is_del", False), _connector="AND")
        q2 = Q(("msg_title__contains", '了你在星圈发布的帖子'), ("source", 1), _connector="OR")

        now = timezone.now()
        q1.children.append(("create_time__lte", now))
        q1.children.append(("create_time__gt", now - timezone.timedelta(days=90)))

        app_id = params.get('app_id')
        if app_id:
            app_args = ('app_id', app_id)
            q1.children.append(app_args)

        msg_type = params.get('msg_type')
        if msg_type:
            type_args = ('msg_type', msg_type)
            q1.children.append(type_args)

        msg_title = params.get('msg_title')
        if msg_title:
            title_args = ('msg_title__icontains', msg_title)
            q1.children.append(title_args)

        return DingMessageModel.objects.filter(~q2, q1).order_by('-id').all()

    def get_unique_message_list(self, params=None):
        """ 获取消息 """
        msg_set = set()
        ordered_message_map = OrderedDict()
        msg_queryset = self.get_message_queryset(params=params).values("id", "msg_title", "app_id")[:200]

        for msg_item in msg_queryset:
            msg_id = msg_item["id"]
            app_id = str(msg_item["app_id"])
            msg_title = msg_item["msg_title"]

            md5 = BaseCipher.crypt_md5(msg_title)
            key = (app_id, md5)
            message_list = ordered_message_map.setdefault(app_id, [])

            if key not in msg_set:
                msg_set.add(key)
                message_list.append(dict(msg_id=msg_id, msg_title=msg_title))

        return ordered_message_map

    def get_queryset(self):
        app_queryset = DingAppTokenModel.objects.filter(is_del=False).values("id", "app_name", "app_token")
        message_map = self.get_unique_message_list()
        source_list = CreateDingMessageView.get_source_list()

        return dict(app_list=list(app_queryset), message_map=message_map, source_list=source_list)


class ListPeriodicTaskView(ListVueView):
    """ 定时任务列表 """
    template_name = "ding_talk/ding_periodic_task_list.html"
    serializer_class = DingPeriodicTaskSerializer

    def get_pagination_list(self):
        queryset = DingPeriodicTaskModel.objects.filter(is_del=False).all()
        periodic_task_list = self.serializer_class(queryset[:10], many=True).data
        periodic_task_list.sort(key=itemgetter('enabled'), reverse=True)

        return dict(list=periodic_task_list, total_count=queryset.count())


class AddMessageInfoView(ListVueView):
    """ 消息列表 """
    template_name = "ding_talk/ding_periodic_topic_add.html"
    serializer_class = DingMessageSerializer

    def get_pagination_list(self):
        queryset = AddPeriodicTaskView.get_message_queryset()
        serializer_data_list = self.serializer_class(queryset[:10], many=True).data

        app_queryset = DingAppTokenModel.objects.filter(is_del=False).values("id", "app_name", "app_token")
        msg_type_list = CreateDingMessageView.get_msg_type_list()  # 目前仅用OA

        # 媒体文件
        app_media_mapping = CreateDingMessageView.get_media_map()
        source_list = CreateDingMessageView.get_source_list()

        return dict(
            list=serializer_data_list, app_list=list(app_queryset),
            media_list=msg_type_list, total_count=queryset.count(),
            app_media_mapping=app_media_mapping, source_list=source_list,
        )


class OperateMessageInfoApi(ListAPIView):
    serializer_class = DingMessageSerializer

    def search(self):
        # 关键字搜索
        query_params = self.request.query_params
        message_map = AddPeriodicTaskView().get_unique_message_list(params=query_params)

        return message_map

    def get_queryset(self):
        return AddPeriodicTaskView.get_message_queryset(params=self.request.query_params)

    def get(self, request, *args, **kwargs):
        """ 关键字搜索 """
        query_params = request.query_params
        action = query_params.get("action", 'list')

        if action == 'list':
            return self.list(request, *args, **kwargs)  # 翻页搜索

        return Response(data=getattr(self, action)())

    def edit(self, *args, **kwargs):
        """ 新增与修改消息 """
        message_id = self.request.data.pop("message_id", 0)
        instance = self.serializer_class.Meta.model.objects.filter(id=message_id).first()

        serializer = self.serializer_class(instance, data=self.request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return serializer.data

    def remove(self, *args, **kwargs):
        message_id = self.request.data.get("id") or 0
        DingMessageModel.objects.filter(id=message_id).update(is_del=True)

    def post(self, request, *args, **kwargs):
        action = request.data.get("action")
        if not action or not hasattr(self, action):
            return Response(data=dict(code=501, message="没有该方法！"))

        result = getattr(self, action)(*args, **kwargs)
        return Response(data=result)


class OperatePeriodicTaskApi(APIView):
    serializer_class = DingPeriodicTaskSerializer
    ding_periodic_model = DingPeriodicTaskModel

    def get_crontabschedule(self, params=None):
        params = params or self.request.data
        get_time = (lambda k: (params.get(k) or "").strip())

        cron_obj, is_create = CrontabSchedule.objects.get_or_create(
            minute=get_time("minute"), hour=get_time("hour"),
            day_of_month=get_time("day_of_month"), month_of_year=get_time("month_of_year"),
            day_of_week=get_time("day_of_week"), timezone=pytz.timezone(settings.TIME_ZONE)
        )

        return cron_obj, is_create

    def get_deadline_run_time(self, cron_expr, max_run_times=0):
        if not max_run_times:
            return

        run_time_list = CronParseApi.get_cron_run_time_list(cron_expr, max_run_times)
        deadline_run_time = datetime.strptime(run_time_list[-1], "%Y-%m-%d %H:%M:%S")

        return deadline_run_time

    def _get_ding_cron_kwargs(self, params=None, crontab_id=None, periodic_task_id=None):
        params = params or self.request.data

        cron_expr = params.get("cron_expr")
        message_id = int(params.get("message_id") or 0)
        max_run_times = int(params.get('max_run_times') or 0)

        # Format: push_range
        # {
        #     "dep_ids": ['c7d6b456-24a5-4a30-973e-9607b0a05071']           # 部门ids
        #     "receiver_mobile": '13600000000,13500000001,13500000002'     # 接收人手机号
        # }
        push_range = params.get('push_range') or {}
        receiver_mobiles = [s.strip() for s in (push_range.get('receiver_mobile') or '').split(',') if s.strip()]

        # excel 批量上传的手机号
        excel_path = params.pop("batch_mobile_path", None)
        receiver_mobiles.extend(SendDingAppMessageApi.get_mobiles_from_excel(excel_path))
        push_range['receiver_mobile'] = ','.join(set(receiver_mobiles))

        kwargs = dict(
            cron_name=params.get("cron_name", ""), cron_expr=cron_expr,
            remark=params.get("remark", ""), max_run_times=max_run_times,
            message_id=message_id, push_range=json.dumps(push_range),
            deadline_run_time=self.get_deadline_run_time(cron_expr, max_run_times),
        )

        if crontab_id:
            kwargs.update(beat_cron_id=crontab_id)

        if periodic_task_id:
            kwargs.update(beat_periodic_task_id=periodic_task_id)

        return kwargs

    def list(self):
        """ 任务列表 """
        query_params = self.request.query_params

        query_kwargs = dict(is_del=False)
        cron_name = query_params.get('cron_name')
        remark = query_params.get('remark')

        cron_name and query_kwargs.update(cron_name__icontains=cron_name)
        remark and query_kwargs.update(remark__icontains=remark)

        page = int(query_params.get('page', 1))
        page_size = int(query_params.get('page_size', 10))
        start = (page - 1) * page_size
        end = start + page_size

        queryset = self.ding_periodic_model.objects.filter(**query_kwargs).all()
        serializer_data_list = self.serializer_class(queryset[start:end], many=True).data
        serializer_data_list.sort(key=itemgetter('enabled'), reverse=True)

        return dict(list=serializer_data_list, page=page, total_count=queryset.count())

    def add(self):
        """ 添加任务 """
        from fosun_circle.apps.ding_talk.tasks.task_periodic_ding_message import send_periodic_ding_message

        params = self.request.data
        msg_kwargs = dict(app_id=params["app_id"], id=params["message_id"], is_del=False)
        msg_object = DingMessageModel.objects.filter(**msg_kwargs).first()

        if not msg_object:
            return dict(code=5003, message="该消息<ID: %s>不存在" % msg_object.id)

        celery_task = send_periodic_ding_message
        cron_obj, is_create = self.get_crontabschedule(params)  # 任务定时时间(Crontab)

        with transaction.atomic():
            crontab_id = cron_obj.id
            ding_cron_kwargs = self._get_ding_cron_kwargs(params, crontab_id=crontab_id)

            try:
                name_abbr = celery_task.__name__ + datetime.now().strftime("%Y%m%d%H%M%S")
                ding_cron_obj = self.ding_periodic_model.objects.create(**ding_cron_kwargs)

                periodic_task_obj = PeriodicTask.objects.create(
                    name=name_abbr,  # 任务名缩写
                    task=params.get("task_name", celery_task.name),   # 完整任务名路径
                    enabled=False, crontab_id=crontab_id,
                    kwargs=json.dumps(dict(ding_cron_id=ding_cron_obj.id)),
                )

                ding_cron_obj.beat_periodic_task_id = periodic_task_obj.id
                ding_cron_obj.save()
            except Exception as e:
                logger.info("创建动态定时任务失败, name: %s, crontab_id: %s, err： %s", name_abbr, crontab_id, e)
                is_create and cron_obj.delete()  # 新增的删掉

    def update(self):
        """ 编辑任务 """
        data = self.request.data
        ding_periodic_id = data.get("ding_periodic_id") or 0
        ding_periodic_obj = self.ding_periodic_model.objects.filter(id=ding_periodic_id, is_del=False).first()

        if ding_periodic_obj:
            cron_obj, is_create = self.get_crontabschedule(data)  # 任务定时时间
            crontab_id = cron_obj.id

            beat_periodic_task_id = ding_periodic_obj.beat_periodic_task_id
            beat_periodic_obj = PeriodicTask.objects.filter(id=beat_periodic_task_id).first()

            try:
                ding_cron_kwargs = self._get_ding_cron_kwargs(data, crontab_id=crontab_id)

                with transaction.atomic():
                    ding_periodic_obj.save_attributes(force_update=True, **ding_cron_kwargs)
                    beat_periodic_obj.crontab_id = crontab_id
                    beat_periodic_obj.save()

                    if beat_periodic_obj.enabled:
                        PeriodicTasks.update_changed()
                return dict(message='编辑任务<ding_periodic_id: %s> OK' % ding_periodic_id)
            except Exception as e:
                logger.error(traceback.format_exc())
                is_create and cron_obj.delete()  # 新增的删掉

                raise ValidationError("编辑定时任务<ding_periodic_id: %s>失败: err: %s" % (ding_periodic_id, e))

        raise ValidationError('编辑任务<ding_periodic_id: %s>不存在。' % ding_periodic_id)

    def delete(self):
        """ 删除任务 """
        ding_periodic_id = self.request.data.get("ding_periodic_id") or 0
        ding_periodic_obj = self.ding_periodic_model.objects.filter(id=ding_periodic_id, is_del=False).first()

        if ding_periodic_obj:
            beat_periodic_task_id = ding_periodic_obj.beat_periodic_task_id
            beat_periodic_obj = PeriodicTask.objects.filter(id=beat_periodic_task_id).first()

            if beat_periodic_obj:
                with transaction.atomic():
                    ding_periodic_obj.save_attributes(force_update=True, is_del=True)
                    beat_periodic_obj.enabled = False
                    beat_periodic_obj.save()

                    PeriodicTasks.update_changed()

                return dict(message='删除任务<ding_periodic_id: %s> OK' % ding_periodic_id)

        raise ValidationError('删除任务<ding_periodic_id: %s>不存在。' % ding_periodic_id)

    def set_enabled(self):
        """ 设置任务状态 """
        ding_periodic_id = self.request.data.get("ding_periodic_id") or 0
        ding_periodic_obj = self.ding_periodic_model.objects.filter(id=ding_periodic_id, is_del=False).first()

        if ding_periodic_obj:
            beat_periodic_task_id = ding_periodic_obj.beat_periodic_task_id
            beat_periodic_obj = PeriodicTask.objects.filter(id=beat_periodic_task_id).first()

            if beat_periodic_obj:
                enabled = not beat_periodic_obj.enabled
                cron_expr = ding_periodic_obj.cron_expr
                max_run_times = ding_periodic_obj.max_run_times

                with transaction.atomic():
                    if enabled:
                        ding_periodic_obj.deadline_run_time = self.get_deadline_run_time(cron_expr, max_run_times)
                        ding_periodic_obj.save()

                    beat_periodic_obj.enabled = enabled
                    beat_periodic_obj.save()

                    PeriodicTasks.update_changed()

                return dict(message=(enabled and "启动" or "暂停") + '任务<ding_periodic_id: %s> OK' % ding_periodic_id)

        raise ValidationError('任务<ding_periodic_id: %s>不存在。' % ding_periodic_id)

    def get(self, request, *args, **kwargs):
        action = request.query_params.get("action")
        return Response(data=getattr(self, action)())

    def post(self, request, *args, **kwargs):
        action = request.data.get("action")
        if not action or not hasattr(self, action):
            return Response(data=dict(code=501, message="没有该方法！"))

        result = getattr(self, action)()
        return Response(data=result)

