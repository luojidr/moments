import os.path
import random

import openpyxl
from django.conf import settings
from django.db import connections, transaction
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.generics import mixins
from rest_framework.response import Response
from rest_framework.generics import ListAPIView, CreateAPIView, RetrieveAPIView, GenericAPIView

from ding_talk.views import SendDingAppMessageApi, RecallDingAppMsgTaskApi
from fosun_circle.libs.log import dj_logger as logger
from fosun_circle.core.views import ListVueView, SingleVueView
from ding_talk.models import DingMsgPushLogModel, DingMessageModel, DingAppTokenModel

from .serializers import LotteryActivitySerializer, ParticipantSerializer, AwardSerializer
from.models import ParticipantModel, AwardModel, ActivityModel


class ListLotteryActivityView(ListVueView):
    template_name = "lottery/lottery_list.html"
    serializer_class = LotteryActivitySerializer

    def get_pagination_list(self):
        queryset = self.serializer_class.Meta.model.objects.filter(is_del=False)
        serializer = self.serializer_class(queryset[:10], many=True)
        return dict(list=serializer.data, total_count=queryset.count())


class DetailLotteryActivityView(SingleVueView):
    template_name = "lottery/lottery_detail.html"
    serializer_class = LotteryActivitySerializer

    def get_queryset(self):
        activity_id = int(self.request.GET.get('id') or 1)
        obj = self.serializer_class.Meta.model.objects.get(id=activity_id, is_del=False)
        return obj.to_dict()


class LotteryActivityApi(RetrieveAPIView):
    serializer_class = LotteryActivitySerializer
    queryset = serializer_class.Meta.model.objects.filter(is_del=False).all()

    def get_object(self):
        activity_id = int(self.request.query_params.get('id') or 0)
        return self.queryset.get(id=activity_id, is_del=False)


class ListLotteryActivityApi(ListAPIView):
    serializer_class = LotteryActivitySerializer

    def get_queryset(self):
        name = self.request.query_params.get('name')
        queryset = self.serializer_class.Meta.model.objects.filter(is_del=False)

        if name:
            queryset = queryset.filter(name__icontains=name).all()

        return queryset.order_by('-id').all()


class CreateUpdateLotteryActivityApi(mixins.CreateModelMixin,
                                     mixins.UpdateModelMixin,
                                     GenericAPIView):
    serializer_class = LotteryActivitySerializer
    queryset = serializer_class.Meta.model.objects.all()

    def get_participant_mobiles(self, mode):
        mobile_list = []
        mode_list = [m[0] for m in self.serializer_class.Meta.model.PARTICIPANT_MODES]
        assert mode in mode_list, "参与名单获取方式错误"

        if mode == 'EXCEL':
            path = self.request.data.pop("lottery_import_path", None)
            if not path or not os.path.exists(path):
                return mobile_list

            wb = openpyxl.load_workbook(path)
            sheet = wb.worksheets[0]
            sheet_values = [str(sheet.cell(row, 1).value or '').strip() for row in range(2, sheet.max_row + 1)]
            mobile_list.extend([cell_val for cell_val in sheet_values if cell_val])
        else:
            required_userids = []
            cursor = connections['bbs_user'].cursor()

            tag_id = self.request.data.get("tag_id", 0)
            deadline = self.request.data.get("deadline")

            cursor.execute("SELECT id FROM users_userinfo WHERE fullname='星小圈'")
            admin_userids = [str(ret[0]) for ret in cursor.fetchall()]

            sql = """
                SELECT id, user_id, is_actual FROM "starCircle_starcircle"
                WHERE id IN (SELECT circle_id FROM "starCircle_circle2tag" WHERE tag_id=%s) 
                AND created_time <=%s AND is_show=true AND is_delete=false
            """
            cursor.execute(sql, (tag_id, deadline))
            db_base_default_ret = cursor.fetchall()

            if self.request.data.get("is_posted", False):
                posted_userids = [str(item[1]) for item in db_base_default_ret if item[2]]
                required_userids.extend(posted_userids)

            # 后期实现: 匿名帖子的评论与点赞不纳入参与资格?
            if self.request.data.get("is_commented", False):
                sql = """
                    SELECT a.circle_id, b.userid FROM "starCircle_circle2tag" a 
                    JOIN "starCircle_starcircle" b ON a.circle_id=b.id 
                    WHERE a.tag_id=%s AND b.is_show=true AND b.is_delete=false AND b.is_actual=false
                """
                commented_userids = []
                required_userids.extend(commented_userids)

            if self.request.data.get("is_liked", False):
                liked_userids = []
                required_userids.extend(liked_userids)

            # 获取手机号
            userid_str = ','.join(list(set(required_userids) - set(admin_userids)))
            cursor.execute('SELECT "phoneNumber" FROM users_userinfo WHERE id IN (%s)' % userid_str)
            mobile_list.extend([ret[0] for ret in cursor.fetchall()])

        return mobile_list

    def save_participant(self, activity_id):
        mode = self.request.data.get("participant_mode")

        if mode:
            data = dict(activity_id=activity_id, mobiles=self.get_participant_mobiles(mode))
            serializer = ParticipantSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save()  # create per time

    def get_object(self):
        activity_id = int(self.request.data.get('id') or 0)
        return self.queryset.get(id=activity_id)

    def post(self, request, *args, **kwargs):
        activity_id = int(request.data.get('id') or 0)

        with transaction.atomic():
            if not activity_id:
                response = self.create(request, *args, **kwargs)
            else:
                # partial: 允许部分字段更新
                response = self.update(request, *args, partial=True, **kwargs)

            self.save_participant(response.data['id'])
        return response


class ListLotteryParticipantApi(ListAPIView):
    serializer_class = ParticipantSerializer

    def get_queryset(self):
        query_params = self.request.query_params
        is_awarded = int(query_params.get('is_awarded', 0))
        query = dict(is_del=False, activity_id=int(query_params.get('activity_id') or 0))

        if is_awarded == 1:
            query['is_awarded'] = True

        Model = self.serializer_class.Meta.model
        queryset = Model.objects.filter(**query).select_related('award').order_by('award__award_type')
        return queryset


class ListLotteryAwardApi(ListAPIView):
    serializer_class = AwardSerializer

    def get_queryset(self):
        query_params = self.request.query_params
        query = dict(is_del=False, activity_id=int(query_params.get('activity_id') or 0))
        queryset = self.serializer_class.Meta.model.objects.filter(**query).all()

        return queryset


class CreateLotteryAwardApi(CreateAPIView):
    serializer_class = AwardSerializer

    def create(self, request, *args, **kwargs):
        # 建议使用: UpdateAPIView, 在序列化器中重写update, 这样不需要在视图类中重写create或update方法
        # serializer = self.get_serializer(instance=obj, data=request.data, many=True), 尽管这个instance起的作用不大
        # 然后在序列化器中重写update方法: 新增或更新; 这样尽可能保证视图类的干净和简要
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)  # Equal: serializer.save()

        return Response(status=status.HTTP_201_CREATED)


class LuckyWinLotteryApi(APIView):
    """ 抽奖 """
    def post(self, request, *args, **kwargs):
        activity_id = int(request.data.get('activity_id') or 0)
        query = dict(activity_id=activity_id, is_del=False)

        queryset = ParticipantModel.objects.filter(**query).all()
        mobile_list = queryset.values_list('mobile', flat=True)

        award_queryset = AwardModel.objects.filter(**query).values('id', 'num')
        award_nums = sum([item['num'] for item in award_queryset])

        if award_nums > queryset.count():
            raise ValueError('奖品数量大于参与人员')

        # 随机抽奖
        lucky_winners = {}
        for award_item in award_queryset:
            award_id = award_item['id']

            for _ in range(award_item['num']):
                mobile = random.choice(mobile_list)

                while mobile in lucky_winners:
                    mobile = random.choice(mobile_list)

                lucky_winners[mobile] = award_id

        queryset.filter(is_awarded=True).update(is_awarded=False, award_id=None)

        with transaction.atomic():
            for win_mobile, award_id in lucky_winners.items():
                queryset.filter(mobile=win_mobile).update(is_awarded=True, award_id=award_id)

            ActivityModel.objects.filter(id=activity_id, is_del=False).update(awarded_num=award_nums)

        # 中奖人员列表
        # awarded_queryset = queryset.filter(is_awarded=True, mobile__in=lucky_winners).all()
        # serializer = ParticipantSerializer(awarded_queryset, many=True)  # 反序列化: data=request.data
        return Response(data=None, status=status.HTTP_200_OK)


class LuckyPushLotteryApi(APIView):
    """ 推送、撤回中奖消息 """
    def post(self, request, *args, **kwargs):
        action = request.data.get('action')
        activity_id = int(request.data.get('activity_id') or 0)
        participant_id = int(request.data.get('participant_id') or 0)

        assert action in ['push', 'recall'], '无法推送或撤回中奖消息'

        query = dict(activity_id=activity_id, is_del=False,  is_awarded=True)
        participant_id and query.update(id=participant_id)  # 单个推送 | 撤回

        if action == 'push':
            self.push(query)
        else:
            self.recall(query)

        return Response(data=None, status=status.HTTP_200_OK)

    def push(self, query):
        """ 推送单个或多个中奖消息  """
        activity_id = query['activity_id']
        queryset = ParticipantModel.objects.filter(**query).all()
        mobile_list = queryset.values_list('mobile', flat=True)

        send_ding_api = SendDingAppMessageApi()
        app_obj = DingAppTokenModel.objects.get(id=1, is_del=False)
        msg_title = '恭喜你获奖啦'
        msg_text = '恭喜你获得本次<{activity_name}>活动抽奖奖品{award_name}！请复星通联系星小圈领取哦，点击查看更多详情。'
        ding_url = '{host}/exerland/pushPrize/{participant_id}'
        partial_data = dict(
            msg_media='@lALPDeREeGVpxNzNAcnNA0M', msg_type=6, source=6,
            app_token=app_obj.app_token
        )

        # 推送后标记
        msg_uid_list = []
        for mobile in mobile_list:
            participant_obj = queryset.filter(mobile=mobile).select_related('activity', 'award').first()

            if participant_obj:
                activity_name = participant_obj.activity.name
                award_name = participant_obj.award.name

                one_ding_url = ding_url.format(host=settings.FRONT_HOST, participant_id=participant_obj.id)
                push_data = dict(
                    partial_data,
                    msg_url=one_ding_url, msg_pc_url=one_ding_url,
                    msg_title=msg_title, receiver_mobile=mobile,
                    msg_text=msg_text.format(activity_name=activity_name, award_name=award_name)
                )
                instance_list = send_ding_api.async_send_messages(data=push_data)
                logger.info("推送中奖消息成功 => instance_list: %s", instance_list)

                if instance_list:
                    log_obj, *_ = instance_list
                    msg_uid_list.append(log_obj.msg_uid)
                    participant_obj.save_attributes(force_update=True, is_pushed=True, push_log_id=log_obj.id)

        # 可能存在推送延迟(暂略)
        if not query.get('participant_id'):
            ActivityModel.objects.filter(id=activity_id, is_del=False).update(is_pushed=True)

    def recall(self, query):
        """ 撤回单个中奖消息  """
        participant_id = query.get('id')
        if not participant_id:
            raise ValueError('参数错误，无法撤回中奖消息')

        queryset = ParticipantModel.objects.filter(**query).all()
        recall_count = queryset.count()

        if recall_count > 1 or recall_count == 0:
            raise ValueError('撤回中奖消息数量错误<RecallCnt: %s>' % recall_count)

        log_obj = DingMsgPushLogModel.objects.get(id=queryset[0].push_log_id, is_del=False)
        ding_msg_obj = DingMessageModel.objects.filter(id=log_obj.ding_msg_id, is_del=False).select_related('app').first()

        self.request.data.clear()
        self.request.data.update(app_token=ding_msg_obj.app.app_token, msg_uid=log_obj.msg_uid)
        RecallDingAppMsgTaskApi().post(self.request)

        queryset.update(is_recall=True)  # 一条


class LuckyWinnerDetailLotteryApi(APIView):
    def get(self, request, *args, **kwargs):
        """ 中奖后H5消息详情 """
        logger.info("LuckyWinnerDetailLotteryApi => user: %s, mobile: %s", request.user, request.user.phone_number)

        participant_id = int(request.query_params.get('participant_id') or 0)
        obj = ParticipantModel.objects.filter(
            id=participant_id, is_del=False,
            mobile=request.user.phone_number,
        ).select_related('award').first()

        if obj is None:
            return Response(data=dict(text='您没有该活动的中奖信息'), status=status.HTTP_200_OK)

        award_name = obj.award.name
        award_level = dict(AwardModel.Award_CHOICES)[obj.award.award_type]
        text = "<span>恭喜你获得本次活动{award_name}奖！</span>" \
               "<br>" \
               "<span>请在复星通联系星小圈领取</span>".format(award_name=award_name)

        return Response(data=dict(text=text), status=status.HTTP_200_OK)


